"""
流动性压力测试及风险缓释 — API 路由
"""
import json
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
import io
from ..services.report_generator import generate_report
from sqlalchemy.orm import Session
from sqlalchemy import func

from ..database import get_db
from ..dependencies import get_current_user
from ..models.liquidity import IalmdG21Gap, IalmdHqlaAsset, IalmdStressVersion
from ..schemas.common import ResponseBase, PageResponse
from ..schemas.liquidity import (
    G21GapCreate, G21GapUpdate, G21GapVO, G21ImportRequest,
    HqlaAssetCreate, HqlaAssetUpdate, HqlaAssetVO, HqlaImportRequest,
    StressVersionCreate, StressVersionUpdate, StressVersionVO,
    ScenarioParamsUpdate, SCENARIO_DEFAULTS,
    VersionCompareRequest,
)

router = APIRouter(prefix="/api/liquidity", tags=["流动性压力测试"])

# ====================================================================
# G21 流动性期限缺口数据 CRUD
# ====================================================================

@router.get("/g21", response_model=PageResponse)
def list_g21(
    report_period: Optional[str] = Query(None, description="报告期筛选"),
    category: Optional[str] = Query(None, description="分类筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """G21数据列表"""
    q = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.is_deleted == 0, IalmdG21Gap.status == 1
    )
    if report_period:
        q = q.filter(IalmdG21Gap.report_period == report_period)
    if category:
        q = q.filter(IalmdG21Gap.category == category)

    total = q.count()
    items = q.order_by(IalmdG21Gap.category, IalmdG21Gap.item_code).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return PageResponse(
        data=[G21GapVO.model_validate(i).model_dump() for i in items],
        total=total, page=page, page_size=page_size,
    )


@router.get("/g21/periods", response_model=ResponseBase)
def list_g21_periods(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取已导入的G21报告期列表"""
    periods = (
        db.query(IalmdG21Gap.report_period)
        .filter(IalmdG21Gap.is_deleted == 0, IalmdG21Gap.status == 1)
        .distinct()
        .order_by(IalmdG21Gap.report_period.desc())
        .all()
    )
    return ResponseBase(data=[p[0] for p in periods])


@router.get("/g21/{g21_id}", response_model=ResponseBase)
def get_g21(
    g21_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """G21数据详情"""
    item = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.id == g21_id, IalmdG21Gap.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    return ResponseBase(data=G21GapVO.model_validate(item).model_dump())


@router.post("/g21", response_model=ResponseBase)
def create_g21(
    data: G21GapCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """手动录入G21数据"""
    obj = IalmdG21Gap(**data.model_dump(), created_by=current_user.get("id"))
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return ResponseBase(data=G21GapVO.model_validate(obj).model_dump(), message="录入成功")


@router.put("/g21/{g21_id}", response_model=ResponseBase)
def update_g21(
    g21_id: int,
    data: G21GapUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """更新G21数据"""
    item = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.id == g21_id, IalmdG21Gap.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    item.updated_by = current_user.get("id")
    db.commit()
    db.refresh(item)
    return ResponseBase(data=G21GapVO.model_validate(item).model_dump(), message="更新成功")


@router.delete("/g21/{g21_id}", response_model=ResponseBase)
def delete_g21(
    g21_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """软删除G21数据"""
    item = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.id == g21_id, IalmdG21Gap.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    item.is_deleted = 1
    item.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="删除成功")


@router.post("/g21/import", response_model=ResponseBase)
def import_g21(
    data: G21ImportRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """批量导入G21数据"""
    # 先删除该报告期的旧数据
    db.query(IalmdG21Gap).filter(
        IalmdG21Gap.report_period == data.report_period,
        IalmdG21Gap.is_deleted == 0,
    ).update({"is_deleted": 1})

    count = 0
    for item_data in data.items:
        obj = IalmdG21Gap(
            **item_data.model_dump(),
            report_period=data.report_period,
            created_by=current_user.get("id"),
        )
        db.add(obj)
        count += 1

    db.commit()
    return ResponseBase(data={"imported": count}, message=f"成功导入 {count} 条G21数据")


@router.delete("/g21/period/{period}", response_model=ResponseBase)
def delete_g21_period(
    period: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除整个报告期的G21数据"""
    cnt = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.report_period == period, IalmdG21Gap.is_deleted == 0
    ).update({"is_deleted": 1, "updated_by": current_user.get("id")})
    db.commit()
    return ResponseBase(data={"deleted": cnt}, message=f"已删除报告期 {period} 的 {cnt} 条数据")


# ====================================================================
# HQLA 优质流动性资产 CRUD
# ====================================================================

@router.get("/hqla", response_model=PageResponse)
def list_hqla(
    report_period: Optional[str] = Query(None),
    asset_level: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """HQLA资产列表"""
    q = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1
    )
    if report_period:
        q = q.filter(IalmdHqlaAsset.report_period == report_period)
    if asset_level:
        q = q.filter(IalmdHqlaAsset.asset_level == asset_level)

    total = q.count()
    items = q.order_by(IalmdHqlaAsset.asset_level, IalmdHqlaAsset.id).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return PageResponse(
        data=[HqlaAssetVO.model_validate(i).model_dump() for i in items],
        total=total, page=page, page_size=page_size,
    )


@router.get("/hqla/summary", response_model=ResponseBase)
def hqla_summary(
    report_period: str = Query(..., description="报告期"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """HQLA汇总统计"""
    items = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == report_period,
        IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1,
    ).all()

    summary = {"total_hqla": 0, "level1_total": 0, "level2a_total": 0, "level2b_total": 0,
               "total_face_value": 0, "level1_ratio": 0, "level2_ratio": 0, "level2b_ratio": 0}
    for item in items:
        v = float(item.hqla_value or 0)
        summary["total_hqla"] += v
        summary["total_face_value"] += float(item.face_value or 0)
        if item.asset_level == "LEVEL1":
            summary["level1_total"] += v
        elif item.asset_level == "LEVEL2A":
            summary["level2a_total"] += v
        elif item.asset_level == "LEVEL2B":
            summary["level2b_total"] += v

    if summary["total_hqla"] > 0:
        summary["level1_ratio"] = round(summary["level1_total"] / summary["total_hqla"] * 100, 1)
        summary["level2_ratio"] = round((summary["level2a_total"] + summary["level2b_total"]) / summary["total_hqla"] * 100, 2)
        summary["level2b_ratio"] = round(summary["level2b_total"] / summary["total_hqla"] * 100, 2)

    summary["compliance"] = {
        "level2_limit_ok": summary["level2_ratio"] <= 40,
        "level2b_limit_ok": summary["level2b_ratio"] <= 15,
        "level1_min_ok": summary["level1_ratio"] >= 60,
    }
    summary["count"] = len(items)
    return ResponseBase(data=summary)


@router.get("/hqla/periods", response_model=ResponseBase)
def list_hqla_periods(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取已维护的HQLA快照期列表"""
    periods = (
        db.query(IalmdHqlaAsset.report_period)
        .filter(IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1)
        .distinct()
        .order_by(IalmdHqlaAsset.report_period.desc())
        .all()
    )
    return ResponseBase(data=[p[0] for p in periods])


@router.get("/hqla/{asset_id}", response_model=ResponseBase)
def get_hqla(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """HQLA资产详情"""
    item = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.id == asset_id, IalmdHqlaAsset.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    return ResponseBase(data=HqlaAssetVO.model_validate(item).model_dump())


@router.post("/hqla", response_model=ResponseBase)
def create_hqla(
    data: HqlaAssetCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """新增HQLA资产"""
    d = data.model_dump()
    if d["discounted_value"] == 0 and d["market_value"] > 0:
        d["discounted_value"] = d["market_value"] * (1 - d["haircut_rate"])
    if d["hqla_value"] == 0:
        d["hqla_value"] = d["discounted_value"]
    obj = IalmdHqlaAsset(**d, created_by=current_user.get("id"))
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return ResponseBase(data=HqlaAssetVO.model_validate(obj).model_dump(), message="新增成功")


@router.put("/hqla/{asset_id}", response_model=ResponseBase)
def update_hqla(
    asset_id: int,
    data: HqlaAssetUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """更新HQLA资产"""
    item = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.id == asset_id, IalmdHqlaAsset.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    item.updated_by = current_user.get("id")
    db.commit()
    db.refresh(item)
    return ResponseBase(data=HqlaAssetVO.model_validate(item).model_dump(), message="更新成功")


@router.delete("/hqla/{asset_id}", response_model=ResponseBase)
def delete_hqla(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除HQLA资产"""
    item = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.id == asset_id, IalmdHqlaAsset.is_deleted == 0
    ).first()
    if not item:
        return ResponseBase(code=404, message="数据不存在")
    item.is_deleted = 1
    item.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="删除成功")


@router.post("/hqla/import", response_model=ResponseBase)
def import_hqla(
    data: HqlaImportRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """批量导入HQLA资产"""
    db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == data.report_period,
        IalmdHqlaAsset.is_deleted == 0,
    ).update({"is_deleted": 1})

    count = 0
    for item_data in data.items:
        d = item_data.model_dump()
        d["report_period"] = data.report_period
        if d["discounted_value"] == 0 and d["market_value"] > 0:
            d["discounted_value"] = d["market_value"] * (1 - d["haircut_rate"])
        if d["hqla_value"] == 0:
            d["hqla_value"] = d["discounted_value"]
        obj = IalmdHqlaAsset(**d, created_by=current_user.get("id"))
        db.add(obj)
        count += 1

    db.commit()
    return ResponseBase(data={"imported": count}, message=f"成功导入 {count} 条HQLA资产")


@router.delete("/hqla/period/{period}", response_model=ResponseBase)
def delete_hqla_period(
    period: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除整个报告期的HQLA数据"""
    cnt = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == period, IalmdHqlaAsset.is_deleted == 0
    ).update({"is_deleted": 1, "updated_by": current_user.get("id")})
    db.commit()
    return ResponseBase(data={"deleted": cnt}, message=f"已删除报告期 {period} 的 {cnt} 条资产")


# ====================================================================
# Excel 导出
# ====================================================================

@router.get("/hqla/export")
def export_hqla(
    report_period: str = Query(...),
    db: Session = Depends(get_db),
):
    """导出HQLA资产为Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    items = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == report_period,
        IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1,
    ).order_by(IalmdHqlaAsset.asset_level, IalmdHqlaAsset.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HQLA资产明细"

    headers = ["资产层级", "资产名称", "资产类型", "面值(万元)", "市场价值(万元)", "扣减率", "折后价值(万元)", "计入HQLA(万元)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    level_map = {"LEVEL1": "一级资产", "LEVEL2A": "二级A资产", "LEVEL2B": "二级B资产"}
    for i, item in enumerate(items, 2):
        row_data = [
            level_map.get(item.asset_level, item.asset_level),
            item.asset_name, item.asset_type,
            float(item.face_value or 0), float(item.market_value or 0),
            f"{float(item.haircut_rate or 0) * 100:.0f}%",
            float(item.discounted_value or 0), float(item.hqla_value or 0),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col >= 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=HQLA_{report_period}.xlsx"},
    )


@router.get("/g21/export")
def export_g21(
    report_period: str = Query(...),
    db: Session = Depends(get_db),
):
    """导出G21数据为Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    items = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.report_period == report_period,
        IalmdG21Gap.is_deleted == 0, IalmdG21Gap.status == 1,
    ).order_by(IalmdG21Gap.category, IalmdG21Gap.item_code).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G21期限缺口"

    headers = ["分类", "科目编码", "科目名称", "隔夜", "7天", "14天", "1个月", "3个月", "6个月", "1年", "5年以上", "无期限", "合计"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    cat_map = {"ASSET": "资产端", "LIABILITY": "负债端", "OFF_BALANCE": "表外"}
    for i, item in enumerate(items, 2):
        row_data = [
            cat_map.get(item.category, item.category), item.item_code, item.item_name,
            float(item.overnight_amount or 0), float(item.day7_amount or 0),
            float(item.day14_amount or 0), float(item.month1_amount or 0),
            float(item.month3_amount or 0), float(item.month6_amount or 0),
            float(item.year1_amount or 0), float(item.year5_amount or 0),
            float(item.unlimited_amount or 0), float(item.total_amount or 0),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col >= 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=G21_{report_period}.xlsx"},
    )


# ====================================================================
# 压力测试版本 CRUD
# ====================================================================

@router.get("/versions", response_model=PageResponse)
def list_versions(
    version_status: Optional[str] = Query(None, description="版本状态筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """版本列表"""
    q = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.is_deleted == 0, IalmdStressVersion.status == 1
    )
    if version_status:
        q = q.filter(IalmdStressVersion.version_status == version_status)

    total = q.count()
    items = q.order_by(IalmdStressVersion.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return PageResponse(
        data=[StressVersionVO.model_validate(i).model_dump() for i in items],
        total=total, page=page, page_size=page_size,
    )


@router.get("/versions/{version_id}", response_model=ResponseBase)
def get_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """版本详情"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    return ResponseBase(data=StressVersionVO.model_validate(ver).model_dump())


@router.post("/versions", response_model=ResponseBase)
def create_version(
    data: StressVersionCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """创建版本（自动包含全部4种情景的默认参数）"""
    d = data.model_dump()
    if not d.get("scenario_params_json"):
        d["scenario_params_json"] = dict(SCENARIO_DEFAULTS)
    obj = IalmdStressVersion(**d, created_by=current_user.get("id"))
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return ResponseBase(data=StressVersionVO.model_validate(obj).model_dump(), message="版本创建成功")


@router.put("/versions/{version_id}", response_model=ResponseBase)
def update_version(
    version_id: int,
    data: StressVersionUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """更新版本"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    if ver.version_status == "PUBLISHED":
        return ResponseBase(code=400, message="已发布版本不可修改")

    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(ver, k, v)
    ver.updated_by = current_user.get("id")
    db.commit()
    db.refresh(ver)
    return ResponseBase(data=StressVersionVO.model_validate(ver).model_dump(), message="更新成功")


@router.put("/versions/{version_id}/scenario-params", response_model=ResponseBase)
def update_scenario_params(
    version_id: int,
    data: ScenarioParamsUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """更新版本中某个情景的参数"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")

    params = ver.scenario_params_json or {}
    params[data.scenario_type] = data.params
    ver.scenario_params_json = params
    ver.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(data={"scenario_params_json": params}, message=f"{data.scenario_type} 参数已更新")


@router.put("/versions/{version_id}/publish", response_model=ResponseBase)
def publish_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """发布版本"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    ver.version_status = "PUBLISHED"
    ver.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="版本已发布")


@router.put("/versions/{version_id}/recall", response_model=ResponseBase)
def recall_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """收回已发布版本，恢复为草稿状态"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    if ver.version_status != "PUBLISHED":
        return ResponseBase(code=400, message="仅已发布版本可收回")
    ver.version_status = "DRAFT"
    ver.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="版本已收回到草稿状态，参数可重新编辑")


@router.put("/versions/{version_id}/archive", response_model=ResponseBase)
def archive_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """归档版本"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    ver.version_status = "ARCHIVED"
    ver.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="版本已归档")


@router.post("/versions/{version_id}/copy", response_model=ResponseBase)
def copy_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """复制版本"""
    src = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not src:
        return ResponseBase(code=404, message="源版本不存在")

    # 生成新版本编号
    max_code = db.query(func.max(IalmdStressVersion.id)).scalar() or 0
    new_code = f"COPY-{max_code + 1}"

    new_ver = IalmdStressVersion(
        version_code=new_code,
        version_name=f"{src.version_name}（副本）",
        version_desc=src.version_desc,
        version_status="DRAFT",
        g21_period=src.g21_period,
        hqla_period=src.hqla_period,
        test_window=src.test_window,
        scenario_params_json=src.scenario_params_json,
        created_by=current_user.get("id"),
    )
    db.add(new_ver)
    db.commit()
    db.refresh(new_ver)
    return ResponseBase(data=StressVersionVO.model_validate(new_ver).model_dump(), message="版本复制成功")


@router.delete("/versions/{version_id}", response_model=ResponseBase)
def delete_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除版本（仅草稿可删）"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")
    if ver.version_status == "PUBLISHED":
        return ResponseBase(code=400, message="已发布版本不可删除，请先归档")
    ver.is_deleted = 1
    ver.updated_by = current_user.get("id")
    db.commit()
    return ResponseBase(message="版本已删除")


# ====================================================================
# 版本对比
# ====================================================================

@router.post("/versions/compare", response_model=ResponseBase)
def compare_versions(
    data: VersionCompareRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """版本对比"""
    if len(data.version_ids) < 2 or len(data.version_ids) > 4:
        return ResponseBase(code=400, message="请选择2~4个版本进行对比")

    versions = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id.in_(data.version_ids),
        IalmdStressVersion.is_deleted == 0,
    ).order_by(IalmdStressVersion.created_at).all()

    if len(versions) < 2:
        return ResponseBase(code=400, message="未找到足够版本")

    compare_data = []
    for v in versions:
        compare_data.append({
            "id": v.id,
            "version_code": v.version_code,
            "version_name": v.version_name,
            "version_status": v.version_status,
            "test_window": v.test_window,
            "g21_period": v.g21_period,
            "hqla_period": v.hqla_period,
            "scenario_params": v.scenario_params_json,
            "stress_results": v.stress_results_json,
            "cash_flow_gaps": v.cash_flow_gaps_json,
            "mitigation_measures": v.mitigation_measures_json,
            "mitigation_results": v.mitigation_results_json,
            "created_at": v.created_at.isoformat() if v.created_at else None,
        })

    return ResponseBase(data={"versions": compare_data, "count": len(compare_data)})


# ====================================================================
# 压力测试执行（计算引擎）
# ====================================================================

@router.post("/versions/{version_id}/run", response_model=ResponseBase)
def run_stress_test(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """执行压力测试（运行全部4种情景：基准/轻度/中度/重度）"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")

    g21_items = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.report_period == ver.g21_period,
        IalmdG21Gap.is_deleted == 0, IalmdG21Gap.status == 1,
    ).all()

    hqla_items = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == ver.hqla_period,
        IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1,
    ).all()

    total_hqla = sum(float(a.hqla_value or 0) for a in hqla_items)
    total_asset_inflow_30d = sum(
        float(i.overnight_amount or 0) + float(i.day7_amount or 0) + float(i.day14_amount or 0) + float(i.month1_amount or 0)
        for i in g21_items if i.category == "ASSET"
    )
    total_liability_outflow_30d = sum(
        float(i.overnight_amount or 0) + float(i.day7_amount or 0) + float(i.day14_amount or 0) + float(i.month1_amount or 0)
        for i in g21_items if i.category == "LIABILITY"
    )

    scenario_params = ver.scenario_params_json or SCENARIO_DEFAULTS
    all_results = {}
    all_cash_flows = {}

    for scenario, params in scenario_params.items():
        runoff_retail = float(params.get("deposit_runoff_retail", 0))
        runoff_corp = float(params.get("deposit_runoff_corp", 0))
        rollover_rate = float(params.get("wholesale_rollover_rate", 1.0))
        drawdown_rate = float(params.get("credit_drawdown_rate", 0))
        bond_haircut = float(params.get("bond_haircut", 0))
        spread_bp = float(params.get("interbank_spread_bp", 0))

        stressed_outflow = total_liability_outflow_30d * (1 + runoff_retail * 0.6 + runoff_corp * 0.4 + drawdown_rate * 0.3)
        stressed_inflow = total_asset_inflow_30d * (1 - runoff_retail * 0.2 - runoff_corp * 0.15) * rollover_rate * 0.75
        stressed_net_outflow = max(stressed_outflow - stressed_inflow, 1)
        stressed_hqla = total_hqla * (1 - bond_haircut * 0.5 - spread_bp / 10000)

        lcr = (stressed_hqla / stressed_net_outflow * 100) if stressed_net_outflow > 0 else 0
        asf = total_hqla * (1 - runoff_corp * 0.3)
        rsf = total_liability_outflow_30d * (1 + runoff_corp * 0.5)
        nsfr = (asf / rsf * 100) if rsf > 0 else 0
        survival = min(30, int(stressed_hqla / (stressed_outflow / 30))) if stressed_outflow > 0 else 30
        hqla_consumption = round((total_hqla - stressed_hqla) / total_hqla * 100, 1) if total_hqla > 0 else 0

        all_results[scenario] = {
            "lcr": round(lcr, 1),
            "nsfr": round(nsfr, 1),
            "cash_flow_gap": round(stressed_inflow - stressed_outflow, 1),
            "hqla_consumption_rate": hqla_consumption,
            "survival_days": survival,
        }

        periods = ["overnight", "day7", "day14", "month1", "month3", "month6", "year1", "year5"]
        gaps = []
        cumulative = 0
        for p in periods:
            asset_v = sum(float(getattr(i, f"{p}_amount", 0) or 0) for i in g21_items if i.category == "ASSET")
            liability_v = sum(float(getattr(i, f"{p}_amount", 0) or 0) for i in g21_items if i.category == "LIABILITY")
            adj_a = asset_v * (1 - runoff_retail * 0.15 - runoff_corp * 0.1)
            adj_l = liability_v * (1 + runoff_retail * 0.6 + runoff_corp * 0.4)
            gap = round(adj_a - adj_l, 1)
            cumulative += gap
            gaps.append({"period": p, "adj_asset": round(adj_a, 1), "adj_liability": round(adj_l, 1),
                         "net_gap": gap, "cumulative_gap": round(cumulative, 1)})
        all_cash_flows[scenario] = gaps

    ver.stress_results_json = all_results
    ver.cash_flow_gaps_json = all_cash_flows
    ver.updated_by = current_user.get("id")
    db.commit()

    return ResponseBase(data={
        "results": all_results,
        "cash_flow_gaps": all_cash_flows,
    }, message="全部4种情景压力测试执行完成")


@router.get("/versions/{version_id}/report")
def download_report(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """生成并下载流动性风险压力测试报告(docx)"""
    ver = db.query(IalmdStressVersion).filter(
        IalmdStressVersion.id == version_id, IalmdStressVersion.is_deleted == 0
    ).first()
    if not ver:
        return ResponseBase(code=404, message="版本不存在")

    g21_items = db.query(IalmdG21Gap).filter(
        IalmdG21Gap.report_period == ver.g21_period,
        IalmdG21Gap.is_deleted == 0, IalmdG21Gap.status == 1,
    ).all()
    hqla_items = db.query(IalmdHqlaAsset).filter(
        IalmdHqlaAsset.report_period == ver.hqla_period,
        IalmdHqlaAsset.is_deleted == 0, IalmdHqlaAsset.status == 1,
    ).all()

    g21_data = [{c.name: getattr(r, c.name) for c in r.__table__.columns} for r in g21_items]
    hqla_data = [{c.name: getattr(r, c.name) for c in r.__table__.columns} for r in hqla_items]

    version_info = {
        "version_code": ver.version_code,
        "version_name": ver.version_name,
        "g21_period": ver.g21_period,
        "hqla_period": ver.hqla_period,
        "test_window": ver.test_window,
    }

    report_bytes = generate_report(
        version_info, g21_data, hqla_data,
        ver.stress_results_json or {},
        ver.scenario_params_json or {},
        ver.cash_flow_gaps_json or {},
    )

    filename = f"流动性压力测试报告_{ver.version_code}.docx"
    return StreamingResponse(
        io.BytesIO(report_bytes),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={filename.encode('utf-8').decode('latin-1')}"},
    )
